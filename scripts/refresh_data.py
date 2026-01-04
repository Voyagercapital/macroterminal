import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from statistics import mean, pstdev
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
INDICATORS_PATH = os.path.join(os.path.dirname(__file__), 'indicators.json')
OUT_PATH = os.path.join(ROOT, 'public', 'data', 'dashboard.json')

FRED_BASE = 'https://api.stlouisfed.org/fred/series/observations'

@dataclass
class Observation:
    date: str
    value: float


def _safe_float(x: Any) -> Optional[float]:
    try:
        if x is None:
            return None
        if isinstance(x, (int, float)):
            return float(x)
        s = str(x).strip()
        if s in ('', '.', 'NaN', 'nan', 'null', 'None'):
            return None
        return float(s)
    except Exception:
        return None


def zscore(values: List[float], x: float) -> Optional[float]:
    if len(values) < 10:
        return None
    m = mean(values)
    sd = pstdev(values)
    if sd == 0:
        return None
    return (x - m) / sd


def signal_from_z(z: Optional[float], high_is_risk_off: bool) -> str:
    if z is None:
        return 'neutral'

    # Symmetric buckets; direction determined by high_is_risk_off.
    if abs(z) < 0.5:
        return 'neutral'

    if high_is_risk_off:
        if z > 1.5:
            return 'red'
        if z > 0.5:
            return 'amber'
        if z < -1.5:
            return 'green'
        if z < -0.5:
            return 'green'
        return 'neutral'
    else:
        # Low values are risk-off (e.g., jobless claims can be risk-off when low)
        if z < -1.5:
            return 'red'
        if z < -0.5:
            return 'amber'
        if z > 0.5:
            return 'green'
        return 'neutral'


def fetch_fred_series(series_id: str, api_key: str, limit: int = 4000) -> List[Observation]:
    params = {
        'series_id': series_id,
        'api_key': api_key,
        'file_type': 'json',
        'sort_order': 'asc',
        'limit': str(limit),
    }
    r = requests.get(FRED_BASE, params=params, timeout=30)
    r.raise_for_status()
    payload = r.json()
    obs: List[Observation] = []
    for o in payload.get('observations', []):
        v = _safe_float(o.get('value'))
        if v is None:
            continue
        obs.append(Observation(date=o.get('date', ''), value=v))
    return obs


def fetch_rbnz_xlsx(url: str, sheet: int = 0, date_col: str = 'Date', value_col: str = 'TWI') -> List[Observation]:
    '''Fetch a single series column from an RBNZ xlsx file.

    The RBNZ publishes many statistical series as spreadsheets. This function downloads
    the file and extracts a date column and one value column by matching headers.
    '''
    from io import BytesIO

    r = requests.get(url, timeout=45)
    r.raise_for_status()
    wb = load_workbook(BytesIO(r.content), data_only=True)

    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]

    header_row = None
    date_idx = None
    value_idx = None

    def _match(cell, needle: str) -> bool:
        return isinstance(cell, str) and needle.lower() in cell.lower()

    # Find header row within first 30 rows
    for row in range(1, min(31, ws.max_row) + 1):
        vals = [ws.cell(row=row, column=c).value for c in range(1, min(80, ws.max_column) + 1)]
        for c, v in enumerate(vals, start=1):
            if _match(v, date_col):
                date_idx = c
            if _match(v, value_col):
                value_idx = c
        if date_idx is not None and value_idx is not None:
            header_row = row
            break

    if header_row is None or date_idx is None or value_idx is None:
        raise RuntimeError(f"Could not find headers '{date_col}' and '{value_col}' in {url}")

    obs: List[Observation] = []
    for row in range(header_row + 1, ws.max_row + 1):
        dt = ws.cell(row=row, column=date_idx).value
        val = ws.cell(row=row, column=value_idx).value
        v = _safe_float(val)
        if v is None or dt is None:
            continue

        if isinstance(dt, datetime):
            date_str = dt.date().isoformat()
        else:
            date_str = str(dt)
        obs.append(Observation(date=date_str, value=v))

    obs.sort(key=lambda x: x.date)
    return obs



def compute_indicator(ind: Dict[str, Any], fred_key: Optional[str]) -> Tuple[Optional[float], Optional[float], Optional[float], str]:
    source = ind['source']
    series: List[Observation]

    try:
        if source == 'fred':
            if not fred_key:
                return None, None, None, 'neutral'
            series = fetch_fred_series(ind['series_id'], fred_key)
        elif source == 'rbnz_xlsx':
            series = fetch_rbnz_xlsx(
                ind['url'],
                sheet=int(ind.get('sheet', 0)),
                date_col=str(ind.get('date_col', 'Date')),
                value_col=str(ind.get('value_col', '')),
            )
        else:
            # Stub for future sources (ECB/IMF/OECD/BIS/StatsNZ)
            return None, None, None, 'neutral'
    except Exception:
        return None, None, None, 'neutral'

    if len(series) < 2:
        return None, None, None, 'neutral'

    latest = series[-1].value
    prev = series[-2].value
    delta = latest - prev

    # zscore from the last 252 points (or fewer)
    window = [o.value for o in series[-252:]]
    z = zscore(window, latest)

    sig = signal_from_z(z, bool(ind.get('high_is_risk_off', True)))
    return latest, delta, z, sig



def score_from_signal(sig: str) -> int:
    return {'green': -1, 'neutral': 0, 'amber': 1, 'red': 2}.get(sig, 0)


def main() -> None:
    with open(INDICATORS_PATH, 'r', encoding='utf-8') as f:
        cfg = json.load(f)

    fred_key = os.getenv('FRED_API_KEY')

    drivers_out = []
    by_region: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}

    total_weight = 0.0
    total_score = 0.0

    for ind in cfg['drivers']:
        latest, delta, z, sig = compute_indicator(ind, fred_key)
        weight = float(ind.get('weight', 1))

        if latest is not None:
            total_weight += weight
            total_score += weight * score_from_signal(sig)

        item = {
            'name': ind['name'],
            'value': round(latest, 4) if isinstance(latest, float) else latest,
            'delta': round(delta, 4) if isinstance(delta, float) else delta,
            'z': round(z, 4) if isinstance(z, float) else z,
            'signal': sig,
            'source': ind['source'],
            'id': ind.get('series_id') or ind.get('url')
        }

        drivers_out.append(item)

        region = ind.get('region', 'Global')
        block = ind.get('block', 'Other')
        by_region.setdefault(region, {}).setdefault(block, []).append(item)

    # Normalize score into 0..100
    # score range is roughly -1..2 per driver; map to 0..100 via a simple transform
    if total_weight > 0:
        avg = total_score / total_weight  # -1..2
        score = int(round((avg + 1) / 3 * 100))
    else:
        score = 50

    label = risk_label(score)

    regions_out = []
    for region, blocks in by_region.items():
        blocks_out = []
        # Basic verdict strings: can be overridden later when you want more sophistication.
        verdict = {
            'cycle': 'Mixed',
            'inflation': 'Mixed',
            'financial': 'Mixed'
        }
        for title, items in blocks.items():
            blocks_out.append({'title': title, 'items': items})
        regions_out.append({
            'id': region.lower().replace(' ', '-'),
            'name': region,
            'verdict': verdict,
            'blocks': blocks_out
        })

    out = {
        'last_updated_utc': datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z'),
        'risk_state': {
            'score': score,
            'label': label,
            'drivers': sorted(drivers_out, key=lambda x: (-score_from_signal(x['signal']), x['name']))[:8]
        },
        'regions': sorted(regions_out, key=lambda r: r['name'])
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f'Wrote {OUT_PATH}')


def risk_label(score: int) -> str:
    if score >= 75:
        return 'Risk-Off'
    if score >= 60:
        return 'Caution'
    if score <= 35:
        return 'Risk-On'
    return 'Neutral'


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr)
        raise
